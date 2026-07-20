from io import BytesIO
import re
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from app.services.amazon.import_file import parse_amazon_url_file
from app.services.amazon import import_file as import_file_service


def test_csv_import_reads_named_url_column_and_excel_encoding():
    content = (
        "名称,Amazon链接\r\n"
        "Bottle,https://www.amazon.com/dp/B000TEST01\r\n"
        "Bag,https://www.amazon.ca/dp/B000TEST02\r\n"
    ).encode("gb18030")

    assert parse_amazon_url_file("products.csv", content) == [
        "https://www.amazon.com/dp/B000TEST01",
        "https://www.amazon.ca/dp/B000TEST02",
    ]


def test_csv_import_accepts_headerless_first_column():
    content = (
        "https://www.amazon.com/dp/B000TEST01,first\n"
        "https://www.amazon.com/dp/B000TEST02,second\n"
    ).encode()

    assert parse_amazon_url_file("products.csv", content) == [
        "https://www.amazon.com/dp/B000TEST01",
        "https://www.amazon.com/dp/B000TEST02",
    ]


def test_xlsx_import_reads_first_nonempty_worksheet():
    workbook = Workbook()
    workbook.active.title = "Empty"
    worksheet = workbook.create_sheet("Products")
    worksheet.append(["SKU", "source_url"])
    worksheet.append(["A", "https://www.amazon.com/dp/B000TEST01"])
    worksheet.append(["B", "https://www.amazon.de/dp/B000TEST02"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    assert parse_amazon_url_file("products.xlsx", content.getvalue()) == [
        "https://www.amazon.com/dp/B000TEST01",
        "https://www.amazon.de/dp/B000TEST02",
    ]


@pytest.mark.parametrize(
    ("filename", "content", "error"),
    [
        ("products.txt", b"https://amazon.com/dp/B000TEST01", "import_file_type_unsupported"),
        ("products.csv", b"", "import_file_empty"),
        ("products.xlsx", b"not a workbook", "import_file_invalid"),
        ("products.csv", b"url\n", "import_file_has_no_urls"),
    ],
)
def test_import_file_rejects_unsupported_or_invalid_content(filename, content, error):
    with pytest.raises(ValueError, match=error):
        parse_amazon_url_file(filename, content)


def test_import_file_rejects_more_than_100_rows():
    content = ("url\n" + "\n".join(
        f"https://amazon.com/dp/B{index:09d}" for index in range(101)
    )).encode()

    with pytest.raises(ValueError, match="import_file_batch_limit_exceeded"):
        parse_amazon_url_file("products.csv", content)


def test_xlsx_import_rejects_excessive_uncompressed_content(monkeypatch):
    content = BytesIO()
    with ZipFile(content, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"x" * 101)
    monkeypatch.setattr(import_file_service, "MAX_XLSX_UNCOMPRESSED_BYTES", 100)

    with pytest.raises(ValueError, match="import_file_too_complex"):
        parse_amazon_url_file("products.xlsx", content.getvalue())


def test_xlsx_import_converts_lazy_xml_parse_failure_to_validation_error():
    workbook = Workbook()
    workbook.active.append(["url"])
    workbook.active.append(["https://www.amazon.com/dp/B000TEST01"])
    original = BytesIO()
    workbook.save(original)
    workbook.close()
    damaged = BytesIO()
    with ZipFile(original) as source, ZipFile(damaged, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            data = source.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = data[: max(1, len(data) // 2)]
            target.writestr(member, data)

    with pytest.raises(ValueError, match="import_file_invalid"):
        parse_amazon_url_file("products.xlsx", damaged.getvalue())


def test_xlsx_import_rejects_declared_dimensions_before_row_iteration():
    workbook = Workbook()
    workbook.active.cell(row=1, column=101, value="outside scan boundary")
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with pytest.raises(ValueError, match="import_file_worksheet_dimensions_exceeded"):
        parse_amazon_url_file("products.xlsx", content.getvalue())


def test_csv_import_converts_oversized_field_to_validation_error():
    content = ("url\n" + "x" * 200_000).encode()

    with pytest.raises(ValueError, match="import_file_invalid"):
        parse_amazon_url_file("products.csv", content)


def test_xlsx_import_accepts_worksheet_without_declared_dimensions():
    workbook = Workbook()
    workbook.active.append(["url"])
    workbook.active.append(["https://www.amazon.com/dp/B000TEST01"])
    original = BytesIO()
    workbook.save(original)
    workbook.close()
    dimensionless = BytesIO()
    with ZipFile(original) as source, ZipFile(dimensionless, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            data = source.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(br"<dimension\s+ref=\"[^\"]+\"\s*/>", b"", data)
            target.writestr(member, data)

    assert parse_amazon_url_file("products.xlsx", dimensionless.getvalue()) == [
        "https://www.amazon.com/dp/B000TEST01"
    ]
