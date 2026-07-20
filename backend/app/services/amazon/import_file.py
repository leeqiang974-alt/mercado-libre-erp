import csv
from collections.abc import Iterable, Sequence
from xml.etree.ElementTree import ParseError
from io import BytesIO, StringIO
from pathlib import Path
import re
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024
MAX_IMPORT_URLS = 100
MAX_IMPORT_REQUEST_BYTES = MAX_IMPORT_FILE_BYTES + 256 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
MAX_XLSX_MEMBERS = 200
MAX_SCANNED_ROWS = 1_000
MAX_SCANNED_COLUMNS = 100
_URL_HEADERS = {
    "url",
    "link",
    "amazonurl",
    "sourceurl",
    "producturl",
    "amazon链接",
    "商品链接",
    "链接",
}


def parse_amazon_url_file(filename: str, content: bytes) -> list[str]:
    if not filename.strip():
        raise ValueError("import_filename_required")
    if not content:
        raise ValueError("import_file_empty")
    if len(content) > MAX_IMPORT_FILE_BYTES:
        raise ValueError("import_file_too_large")

    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _csv_rows(content)
    elif suffix == ".xlsx":
        rows = _xlsx_rows(content)
    else:
        raise ValueError("import_file_type_unsupported")

    urls = _extract_url_column(rows)
    if not urls:
        raise ValueError("import_file_has_no_urls")
    if len(urls) > MAX_IMPORT_URLS:
        raise ValueError("import_file_batch_limit_exceeded")
    return urls


def _csv_rows(content: bytes) -> Iterable[Sequence[object]]:
    text = None
    encodings = ("utf-16", "utf-8-sig", "gb18030") if content.startswith(
        (b"\xff\xfe", b"\xfe\xff")
    ) else ("utf-8-sig", "gb18030")
    for encoding in encodings:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("import_file_encoding_unsupported")
    try:
        return csv.reader(StringIO(text, newline=""))
    except csv.Error as exc:
        raise ValueError("import_file_invalid") from exc


def _xlsx_rows(content: bytes) -> Iterable[Sequence[object]]:
    _validate_xlsx_archive(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise ValueError("import_file_invalid") from exc
    try:
        return _first_nonempty_worksheet_rows(workbook)
    except (BadZipFile, InvalidFileException, OSError, ParseError, ValueError, KeyError) as exc:
        workbook.close()
        raise ValueError("import_file_invalid") from exc


def _extract_url_column(rows: Iterable[Sequence[object]]) -> list[str]:
    urls: list[str] = []
    column_index: int | None = None
    header_checked = False
    try:
        for scanned_rows, row in enumerate(rows, start=1):
            if scanned_rows > MAX_SCANNED_ROWS:
                raise ValueError("import_file_row_scan_limit_exceeded")
            if len(row) > MAX_SCANNED_COLUMNS:
                raise ValueError("import_file_column_limit_exceeded")
            if not _nonempty_values(row):
                continue
            if not header_checked:
                column_index = next(
                    (
                        index
                        for index, value in enumerate(row)
                        if _normalized_header(value) in _URL_HEADERS
                    ),
                    None,
                )
                header_checked = True
                if column_index is not None:
                    continue
                column_index = 0
            if column_index < len(row) and (value := _cell_text(row[column_index])):
                urls.append(value)
                if len(urls) > MAX_IMPORT_URLS:
                    break
    except (BadZipFile, csv.Error, InvalidFileException, OSError, ParseError, KeyError) as exc:
        raise ValueError("import_file_invalid") from exc
    finally:
        close = getattr(rows, "close", None)
        if callable(close):
            close()
    return urls


def _nonempty_values(row: Sequence[object]) -> list[str]:
    return [value for cell in row if (value := _cell_text(cell))]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized_header(value: object) -> str:
    return re.sub(r"[\s_-]+", "", _cell_text(value).lower())


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ValueError("import_file_too_complex")
            if sum(member.file_size for member in members) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ValueError("import_file_too_complex")
    except BadZipFile as exc:
        raise ValueError("import_file_invalid") from exc


def _first_nonempty_worksheet_rows(workbook):
    def rows():
        try:
            for worksheet in workbook.worksheets:
                if (
                    (worksheet.max_row or 0) > MAX_SCANNED_ROWS
                    or (worksheet.max_column or 0) > MAX_SCANNED_COLUMNS
                ):
                    raise ValueError("import_file_worksheet_dimensions_exceeded")
                iterator = worksheet.iter_rows(values_only=True)
                buffered: list[Sequence[object]] = []
                for index, row in enumerate(iterator, start=1):
                    if index > MAX_SCANNED_ROWS:
                        raise ValueError("import_file_row_scan_limit_exceeded")
                    buffered.append(row)
                    if _nonempty_values(row):
                        yield from buffered
                        yield from iterator
                        return
        finally:
            workbook.close()

    return rows()
